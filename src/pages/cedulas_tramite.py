# =========================
# src/pages/cedulas_tramite.py
# Corrección de cédulas + regeneración + reemplazo seguro de SOLO CC*.pdf
# =========================

from __future__ import annotations

import hashlib
import json
import re
import shutil
from datetime import datetime
import io
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.oracle_jdbc import oracle_connect
from src.cobertura_runner import ArchivoLock, LOCK_PATH, ProcesoCoberturaYaEnEjecucion
from src.cobertura_pdf import generar_coberturas_automaticas_desde_mes


PROJECT_ROOT = Path("/data_nuevo/cobertura_integrada")
OUTPUT_ROOT = Path("/data_nuevo/coberturas")
REPO_ROOT = Path("/data_nuevo/repo_grande/data/datos")
LOGS_DIR = PROJECT_ROOT / "logs"
AUDIT_LOG = LOGS_DIR / "cedulas_tramite_audit.jsonl"
BACKUP_CC_ROOT = LOGS_DIR / "backup_cc_reemplazados"
PDF_CC_REGEX = re.compile(r"^CC(?:_\d{2})?\.pdf$", re.IGNORECASE)


def _ahora_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _ahora_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def _is_pdf_cc(path: Path) -> bool:
    return path.is_file() and PDF_CC_REGEX.fullmatch(path.name) is not None

def _json_hash(data: dict) -> str:
    return hashlib.sha256(json.dumps(data, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()

def _auditar_evento(evento: dict) -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    evento = {"ts": _ahora_iso(), **evento}
    with AUDIT_LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(evento, ensure_ascii=False) + "\n")

def _validar_cedula_o_vacio(valor: str, campo: str) -> str:
    valor = str(valor or "").strip()
    if not valor: return ""
    if not valor.isdigit(): raise ValueError(f"{campo} debe contener solo números.")
    if len(valor) < 10 or len(valor) > 13: raise ValueError(f"{campo} debe tener entre 10 y 13 dígitos.")
    return valor

def _validar_tramite(valor: str) -> str:
    valor = str(valor or "").strip()
    if not valor or not valor.isdigit(): raise ValueError("El trámite debe contener solo números.")
    if len(valor) > 30: raise ValueError("El trámite no debe superar 30 dígitos.")
    return valor

def _limpiar_cc_locales_antes_de_regenerar(tramite: str) -> list[str]:
    origen_dir = OUTPUT_ROOT / tramite
    eliminados: list[str] = []
    if not origen_dir.exists(): return eliminados
    for pdf in sorted(origen_dir.iterdir()):
        if _is_pdf_cc(pdf):
            eliminados.append(str(pdf))
            pdf.unlink()
    return eliminados

def _listar_cc_locales_generados(tramite: str) -> list[Path]:
    origen_dir = OUTPUT_ROOT / tramite
    if not origen_dir.exists() or not origen_dir.is_dir(): return []
    return sorted([p for p in origen_dir.iterdir() if _is_pdf_cc(p)])

def _buscar_destinos_tramite(tramite: str) -> list[Path]:
    if not REPO_ROOT.exists(): raise RuntimeError(f"No existe REPO_ROOT: {REPO_ROOT}")
    repo_root = REPO_ROOT.resolve()
    destinos = []
    for p in REPO_ROOT.rglob(tramite):
        if not p.is_dir(): continue
        if p.name != tramite: continue
        resolved = p.resolve()
        if str(resolved).startswith(str(repo_root)): destinos.append(resolved)
    return sorted(destinos)

def _backup_y_reemplazar_solo_cc(destino_dir: Path, tramite: str, nuevos_pdfs: list[Path], usuario: str) -> dict:
    destino_dir = destino_dir.resolve()
    repo_root = REPO_ROOT.resolve()
    if not str(destino_dir).startswith(str(repo_root)): raise RuntimeError(f"Destino fuera del repositorio oficial: {destino_dir}")
    if not destino_dir.exists() or not destino_dir.is_dir(): raise RuntimeError(f"No existe carpeta destino: {destino_dir}")
    if not nuevos_pdfs: raise RuntimeError("No hay PDFs CC*.pdf nuevos")
    for src_pdf in nuevos_pdfs:
        if not _is_pdf_cc(src_pdf): raise RuntimeError(f"Archivo no permitido: {src_pdf.name}")

    run_id = _ahora_id()
    backup_dir = BACKUP_CC_ROOT / tramite / run_id
    staging_dir = backup_dir / "nuevos_verificados"
    backup_dir.mkdir(parents=True, exist_ok=True)
    staging_dir.mkdir(parents=True, exist_ok=True)
    existentes_cc = sorted([p for p in destino_dir.iterdir() if _is_pdf_cc(p)])

    manifest = {"run_id": run_id, "usuario": usuario, "tramite": tramite, "destino_dir": str(destino_dir),
                "backup_dir": str(backup_dir), "staging_dir": str(staging_dir),
                "existentes_respaldados": [], "nuevos_verificados": [], "eliminados_destino": [],
                "copiados_nuevos": [], "restauracion_por_error": [], "otros_archivos_no_tocados": []}
    for item in sorted(destino_dir.iterdir()):
        if item.is_file() and not _is_pdf_cc(item): manifest["otros_archivos_no_tocados"].append(item.name)

    for old_pdf in existentes_cc:
        backup_pdf = backup_dir / old_pdf.name
        shutil.copy2(old_pdf, backup_pdf)
        manifest["existentes_respaldados"].append({"archivo": old_pdf.name, "origen": str(old_pdf), "backup": str(backup_pdf), "sha256": _sha256_file(backup_pdf)})

    for src_pdf in nuevos_pdfs:
        staged_pdf = staging_dir / src_pdf.name
        shutil.copy2(src_pdf, staged_pdf)
        src_hash = _sha256_file(src_pdf)
        staged_hash = _sha256_file(staged_pdf)
        if src_hash != staged_hash: raise RuntimeError(f"Hash no coincide en staging para {src_pdf.name}")
        manifest["nuevos_verificados"].append({"archivo": src_pdf.name, "origen": str(src_pdf), "staging": str(staged_pdf), "sha256": staged_hash})

    try:
        for old_pdf in existentes_cc:
            old_pdf.unlink()
            manifest["eliminados_destino"].append(str(old_pdf))
        for staged_pdf in sorted(staging_dir.iterdir()):
            if not _is_pdf_cc(staged_pdf): continue
            dst_pdf = destino_dir / staged_pdf.name
            shutil.copy2(staged_pdf, dst_pdf)
            staged_hash = _sha256_file(staged_pdf)
            dst_hash = _sha256_file(dst_pdf)
            if staged_hash != dst_hash: raise RuntimeError(f"Hash no coincide luego de copiar {staged_pdf.name}")
            manifest["copiados_nuevos"].append({"archivo": staged_pdf.name, "origen": str(staged_pdf), "destino": str(dst_pdf), "sha256": dst_hash})
    except Exception as exc:
        for current_cc in sorted([p for p in destino_dir.iterdir() if _is_pdf_cc(p)]):
            try: current_cc.unlink()
            except Exception: pass
        for item in manifest["existentes_respaldados"]:
            backup_pdf = Path(item["backup"])
            restore_pdf = destino_dir / backup_pdf.name
            try:
                shutil.copy2(backup_pdf, restore_pdf)
                manifest["restauracion_por_error"].append({"archivo": backup_pdf.name, "backup": str(backup_pdf), "restaurado": str(restore_pdf)})
            except Exception as restore_exc:
                manifest["restauracion_por_error"].append({"archivo": backup_pdf.name, "backup": str(backup_pdf), "error_restaurando": str(restore_exc)})
        manifest_path = backup_dir / "manifest_reemplazo_cc_error.json"
        manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
        raise RuntimeError(f"Falló el reemplazo de CC*.pdf. Se intentó restaurar. Detalle: {exc}") from exc

    manifest_path = backup_dir / "manifest_reemplazo_cc.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"ok": True, "run_id": run_id, "backup_dir": str(backup_dir), "manifest_path": str(manifest_path),
            "eliminados": len(manifest["eliminados_destino"]), "copiados": len(manifest["copiados_nuevos"]),
            "otros_no_tocados": len(manifest["otros_archivos_no_tocados"]), "detalle": manifest}


def _buscar_por_tramite(username: str, password: str, tramite: str) -> list[dict]:
    conn = ps = rs = None
    sql = """SELECT DIG_TRAMITE, DIG_ID_TRAMITE, DIG_CEDULA, DIG_MENOR_EDAD, DIG_DEPENDIENTE_01, DIG_DEPENDIENTE_02, DIG_COBERTURA, DIG_PLANILLADO, FE_PLA_ANIOMES FROM DIGITALIZACION.DIGITALIZACION WHERE TO_CHAR(DIG_TRAMITE) = ? ORDER BY DIG_ID_TRAMITE"""
    rows: list[dict] = []
    try:
        conn = oracle_connect(username, password)
        ps = conn.jconn.prepareStatement(sql)
        ps.setString(1, tramite); ps.setQueryTimeout(30)
        rs = ps.executeQuery()
        while rs.next():
            rows.append({"DIG_TRAMITE": str(rs.getString(1) or "").strip(), "DIG_ID_TRAMITE": str(rs.getString(2) or "").strip(),
                         "DIG_CEDULA": str(rs.getString(3) or "").strip(), "DIG_MENOR_EDAD": str(rs.getString(4) or "").strip(),
                         "DIG_DEPENDIENTE_01": str(rs.getString(5) or "").strip(), "DIG_DEPENDIENTE_02": str(rs.getString(6) or "").strip(),
                         "DIG_COBERTURA": str(rs.getString(7) or "").strip(), "DIG_PLANILLADO": str(rs.getString(8) or "").strip(),
                         "FE_PLA_ANIOMES": str(rs.getString(9) or "").strip()})
        return rows
    finally:
        for obj in (rs, ps, conn):
            if obj:
                try: obj.close()
                except Exception: pass

def _generar_excel_errores_por_fe_pla(username: str, password: str, fe_pla_aniomes: str) -> bytes | None:
    """Genera Excel en memoria con errores de un FE_PLA_ANIOMES."""
    fe_pla = str(fe_pla_aniomes or "").strip()
    if not fe_pla:
        return None

    # 1. Leer todos los errores de los JSONL
    errores: list[dict] = []
    tramites_en_error: set[str] = set()
    for archivo in sorted(LOGS_DIR.glob("cobertura_auto_*_errors.jsonl")):
        with archivo.open("r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea:
                    continue
                try:
                    obj = json.loads(linea)
                except json.JSONDecodeError:
                    continue
                if obj.get("event") != "PDF_GENERATION_ERROR":
                    continue
                dig_tramite = str(obj.get("dig_tramite", "") or "").strip()
                errores.append({**obj, "_dig_tramite": dig_tramite})
                if dig_tramite:
                    tramites_en_error.add(dig_tramite)

    if not errores:
        return None

    # 2. Consultar Oracle en lotes para mapear dig_tramite → fe_pla_aniomes
    tramite_a_fe_pla: dict[str, str] = {}
    tramites_lista = sorted(tramites_en_error)
    batch_size = 500
    conn = None
    try:
        conn = oracle_connect(username, password)
        for i in range(0, len(tramites_lista), batch_size):
            batch = tramites_lista[i:i + batch_size]
            placeholders = ",".join(["?" for _ in batch])
            sql = f"SELECT TO_CHAR(DIG_TRAMITE), FE_PLA_ANIOMES FROM DIGITALIZACION.DIGITALIZACION WHERE TO_CHAR(DIG_TRAMITE) IN ({placeholders})"
            ps = conn.jconn.prepareStatement(sql)
            for idx, t in enumerate(batch, start=1):
                ps.setString(idx, t)
            rs = ps.executeQuery()
            while rs.next():
                t = str(rs.getString(1) or "").strip()
                fp = str(rs.getString(2) or "").strip()
                tramite_a_fe_pla[t] = fp
            rs.close()
            ps.close()
    except Exception:
        pass
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    # 3. Filtrar por FE_PLA_ANIOMES
    filas: list[dict] = []
    for err in errores:
        t = err["_dig_tramite"]
        fp = tramite_a_fe_pla.get(t, "")
        if fp == fe_pla:
            ts = err.get("ts", "")
            try:
                fecha = datetime.fromisoformat(ts).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                fecha = ts
            filas.append({
                "Fecha": fecha,
                "Run ID": err.get("run_id", ""),
                "Trámite": t,
                "Cédula": err.get("cedula", ""),
                "Tipo persona": err.get("tipo_persona", ""),
                "Categoría error": err.get("error_categoria", ""),
                "Causa probable": err.get("causa_probable", ""),
                "Error": str(err.get("error", "") or "")[:250],
                "PDF esperado": err.get("pdf_path", ""),
                "Segundos": err.get("segundos_pdf", ""),
            })

    if not filas:
        return None

    # 4. Generar Excel en memoria
    wb = Workbook()
    ws = wb.active
    ws.title = f"Errores_{fe_pla}"
    cabeceras = ["Fecha", "Run ID", "Trámite", "Cédula", "Tipo persona", "Categoría error", "Causa probable", "Error", "PDF esperado", "Segundos"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    for col_idx, h in enumerate(cabeceras, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, h in enumerate(cabeceras, start=1):
            ws.cell(row=row_idx, column=col_idx, value=fila.get(h, ""))
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 55
    ws.column_dimensions["H"].width = 60
    ws.column_dimensions["I"].width = 45
    ws.column_dimensions["J"].width = 10
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _contar_por_fe_pla_aniomes(username: str, password: str, fe_pla_aniomes: str) -> dict:
    """Cuenta generados/pendientes para un FE_PLA_ANIOMES."""
    fe_pla = str(fe_pla_aniomes or "").strip()
    if not fe_pla: return {"ok": False, "error": "FE_PLA_ANIOMES vacío."}
    conn = ps = rs = None
    try:
        conn = oracle_connect(username, password)
        ps = conn.jconn.prepareStatement(
            "SELECT COUNT(*), SUM(CASE WHEN TRIM(DIG_COBERTURA)='S' THEN 1 ELSE 0 END) "
            "FROM DIGITALIZACION.DIGITALIZACION "
            "WHERE TRIM(DIG_PLANILLADO)='S' AND TRIM(FE_PLA_ANIOMES)=?")
        ps.setString(1, fe_pla)
        rs = ps.executeQuery()
        if rs.next():
            total = rs.getInt(1) or 0
            generados = rs.getInt(2) or 0
        else:
            total = generados = 0
        pendientes = total - generados
        pct = (generados * 100 / total) if total else 0
        return {"ok": True, "fe_pla_aniomes": fe_pla, "total": total, "generados": generados, "pendientes": pendientes, "pct": round(pct, 1)}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    finally:
        for obj in (rs, ps, conn):
            if obj:
                try: obj.close()
                except Exception: pass


def _actualizar_cedulas_y_resetear_cobertura(username: str, password: str, tramite: str, cedula: str, menor_edad: str, dependiente_01: str, dependiente_02: str) -> dict:
    conn = ps = None
    tramite = str(tramite or "").strip()
    if not tramite: return {"ok": False, "affected": 0, "error": "No DIG_TRAMITE."}
    sql = """UPDATE DIGITALIZACION.DIGITALIZACION SET DIG_CEDULA=?, DIG_MENOR_EDAD=?, DIG_DEPENDIENTE_01=?, DIG_DEPENDIENTE_02=?, DIG_COBERTURA='N' WHERE TO_CHAR(DIG_TRAMITE)=? AND TRIM(DIG_PLANILLADO)='S'"""
    params = [cedula, menor_edad, dependiente_01, dependiente_02, tramite]
    try:
        conn = oracle_connect(username, password); conn.jconn.setAutoCommit(False)
        ps = conn.jconn.prepareStatement(sql)
        for idx, val in enumerate(params, start=1): ps.setString(idx, str(val))
        affected = ps.executeUpdate(); conn.jconn.commit()
        return {"ok": True, "affected": int(affected), "error": ""}
    except Exception as exc:
        if conn:
            try: conn.jconn.rollback()
            except Exception: pass
        return {"ok": False, "affected": 0, "error": str(exc)}
    finally:
        for obj in (ps, conn):
            if obj:
                try: obj.close()
                except Exception: pass

def _actualizar_solo_cobertura(username: str, password: str, tramite: str, valor: str) -> dict:
    conn = ps = None
    valor = str(valor or "").strip().upper()
    tramite = str(tramite or "").strip()
    if valor not in {"S", "N"}: return {"ok": False, "affected": 0, "error": "Valor inválido."}
    if not tramite: return {"ok": False, "affected": 0, "error": "DIG_TRAMITE vacío."}
    try:
        conn = oracle_connect(username, password); conn.jconn.setAutoCommit(False)
        ps = conn.jconn.prepareStatement("UPDATE DIGITALIZACION.DIGITALIZACION SET DIG_COBERTURA=? WHERE TO_CHAR(DIG_TRAMITE)=? AND TRIM(DIG_PLANILLADO)='S'")
        ps.setString(1, valor); ps.setString(2, tramite)
        affected = ps.executeUpdate(); conn.jconn.commit()
        return {"ok": True, "affected": int(affected), "error": ""}
    except Exception as exc:
        if conn:
            try: conn.jconn.rollback()
            except Exception: pass
        return {"ok": False, "affected": 0, "error": str(exc)}
    finally:
        for obj in (ps, conn):
            if obj:
                try: obj.close()
                except Exception: pass

def _corregir_regenerar_y_reemplazar(username: str, password: str, fila: dict, nueva_cedula: str, nueva_menor_edad: str, nuevo_dep1: str, nuevo_dep2: str) -> dict:
    tramite = _validar_tramite(fila.get("DIG_TRAMITE", ""))
    dig_id_tramite = str(fila.get("DIG_ID_TRAMITE", "") or "").strip()
    if not tramite: return {"ok": False, "paso": "VALIDACION", "error": "Sin DIG_TRAMITE."}
    if str(fila.get("DIG_PLANILLADO", "")).strip() != "S": return {"ok": False, "paso": "VALIDACION", "error": "No planillado."}

    nueva_cedula = _validar_cedula_o_vacio(nueva_cedula, "DIG_CEDULA")
    nuevo_dep1 = _validar_cedula_o_vacio(nuevo_dep1, "DIG_DEPENDIENTE_01")
    nuevo_dep2 = _validar_cedula_o_vacio(nuevo_dep2, "DIG_DEPENDIENTE_02")
    nueva_menor_edad = str(nueva_menor_edad or "N").strip().upper()
    if nueva_menor_edad not in {"S", "N"}: return {"ok": False, "paso": "VALIDACION", "error": "MENOR_EDAD inválido."}
    if not nueva_cedula: return {"ok": False, "paso": "VALIDACION", "error": "CÉDULA vacía."}

    antes = {"DIG_CEDULA": fila.get("DIG_CEDULA",""), "DIG_MENOR_EDAD": fila.get("DIG_MENOR_EDAD",""),
             "DIG_DEPENDIENTE_01": fila.get("DIG_DEPENDIENTE_01",""), "DIG_DEPENDIENTE_02": fila.get("DIG_DEPENDIENTE_02",""),
             "DIG_COBERTURA": fila.get("DIG_COBERTURA",""), "DIG_PLANILLADO": fila.get("DIG_PLANILLADO","")}
    despues = {"DIG_CEDULA": nueva_cedula, "DIG_MENOR_EDAD": nueva_menor_edad, "DIG_DEPENDIENTE_01": nuevo_dep1,
               "DIG_DEPENDIENTE_02": nuevo_dep2, "DIG_COBERTURA": "N", "DIG_PLANILLADO": fila.get("DIG_PLANILLADO","")}

    _auditar_evento({"evento": "CORRECCION_CEDULAS_INICIO", "usuario": username, "dig_tramite": tramite,
                     "dig_id_tramite": dig_id_tramite, "antes": antes, "despues": despues,
                     "hash_antes": _json_hash(antes), "hash_despues": _json_hash(despues)})

    fe_pla = str(fila.get("FE_PLA_ANIOMES") or "").strip()
    if not fe_pla: return {"ok": False, "paso": "VALIDACION_FE_PLA", "error": "FE_PLA_ANIOMES vacío."}

    try:
        with ArchivoLock(LOCK_PATH):
            update = _actualizar_cedulas_y_resetear_cobertura(username, password, tramite, nueva_cedula, nueva_menor_edad, nuevo_dep1, nuevo_dep2)
            if not update.get("ok") or int(update.get("affected", 0)) <= 0:
                return {"ok": False, "paso": "ORACLE_UPDATE", "error": update.get("error", "Oracle no actualizó."), "oracle": update}

            locales_eliminados = _limpiar_cc_locales_antes_de_regenerar(tramite)

            try:
                gen = generar_coberturas_automaticas_desde_mes(username=username, password=password, fe_pla_aniomes_desde=fe_pla, dig_tramite=tramite, output_dir=str(OUTPUT_ROOT))
            except Exception as exc:
                _actualizar_solo_cobertura(username, password, tramite, "N")
                return {"ok": False, "paso": "GENERACION", "error": str(exc)}

            nuevos_pdfs = _listar_cc_locales_generados(tramite)
            if not nuevos_pdfs:
                _actualizar_solo_cobertura(username, password, tramite, "N")
                return {"ok": False, "paso": "PDF_LOCAL", "error": f"No se generaron PDFs en {OUTPUT_ROOT / tramite}", "generacion": gen}

            destinos = _buscar_destinos_tramite(tramite)
            if len(destinos) == 0:
                _actualizar_solo_cobertura(username, password, tramite, "N")
                _auditar_evento({"evento": "DESTINO_NO_ENCONTRADO", "usuario": username, "dig_tramite": tramite, "dig_id_tramite": dig_id_tramite, "pdfs_locales": [str(p) for p in nuevos_pdfs]})
                return {"ok": False, "paso": "DESTINO_NO_ENCONTRADO", "error": "No se encontró carpeta destino.", "generacion": gen}
            if len(destinos) > 1:
                _actualizar_solo_cobertura(username, password, tramite, "N")
                return {"ok": False, "paso": "DESTINO_AMBIGUO", "error": "Múltiples destinos.", "generacion": gen}

            try:
                reemplazo = _backup_y_reemplazar_solo_cc(destino_dir=destinos[0], tramite=tramite, nuevos_pdfs=nuevos_pdfs, usuario=username)
            except Exception as exc:
                _actualizar_solo_cobertura(username, password, tramite, "N")
                return {"ok": False, "paso": "REEMPLAZO_CC_DESTINO", "error": str(exc), "generacion": gen}

            cobertura_s = _actualizar_solo_cobertura(username, password, tramite, "S")
    except ProcesoCoberturaYaEnEjecucion as exc:
        return {"ok": False, "paso": "LOCK_GENERACION", "error": str(exc)}

    resultado_final = {"ok": True, "paso": "OK", "dig_tramite": tramite, "dig_id_tramite": dig_id_tramite,
                       "locales_eliminados_antes_de_regenerar": locales_eliminados,
                       "pdfs_locales_generados": [str(p) for p in nuevos_pdfs],
                       "destino": str(destinos[0]), "reemplazo": reemplazo, "generacion": gen, "oracle_cobertura_s": cobertura_s}
    _auditar_evento({"evento": "CORRECCION_CEDULAS_OK", "usuario": username, "dig_tramite": tramite, "dig_id_tramite": dig_id_tramite, "resultado": resultado_final})
    return resultado_final


def cedulas_tramite_page():
    st.markdown("""<div class="main-title">Corrección de cédulas</div>
        <div class="main-subtitle">Actualiza cédulas, regenera coberturas y reemplaza únicamente CC*.pdf del trámite.</div>""", unsafe_allow_html=True)
    st.warning("Esta opción modifica Oracle, regenera PDFs y reemplaza SOLO CC*.pdf en el repositorio. No toca otros PDFs ni crea carpetas.", icon="⚠️")

    username = st.session_state.get("oracle_user", "")
    password = st.session_state.get("oracle_password", "")
    if not username or not password: st.error("No hay sesión de Oracle activa."); return

    st.markdown('<div class="simple-card">', unsafe_allow_html=True)
    tramite_input = st.text_input("Número de trámite", value=st.session_state.get("cedulas_tramite_actual", ""), max_chars=30, key="cedulas_tramite_input").strip()

    col_buscar, col_limpiar = st.columns([3, 1])
    with col_buscar: buscar = st.button("Buscar trámite", key="btn_buscar_tramite", use_container_width=True)
    with col_limpiar:
        if st.button("Limpiar", key="btn_limpiar_cedulas", use_container_width=True):
            for key in ["cedulas_tramite_actual", "cedulas_filas_encontradas", "select_fila_editar_cedulas"]:
                st.session_state.pop(key, None)
            st.rerun()

    if buscar:
        try:
            tramite = _validar_tramite(tramite_input)
            st.session_state["cedulas_tramite_actual"] = tramite
            with st.spinner("Consultando Oracle..."):
                st.session_state["cedulas_filas_encontradas"] = _buscar_por_tramite(username, password, tramite)
        except Exception as exc: st.error(str(exc)); st.markdown("</div>", unsafe_allow_html=True); return

    filas = st.session_state.get("cedulas_filas_encontradas", [])
    if not filas: st.info("Presionar Buscar trámite para consultar Oracle."); st.markdown("</div>", unsafe_allow_html=True); return

    tramite_actual = st.session_state.get("cedulas_tramite_actual", tramite_input)
    st.success(f"Se encontraron {len(filas)} registro(s) para el trámite {tramite_actual}.")
    st.dataframe(pd.DataFrame(filas)[["DIG_ID_TRAMITE","DIG_CEDULA","DIG_MENOR_EDAD","DIG_DEPENDIENTE_01","DIG_DEPENDIENTE_02","DIG_COBERTURA","DIG_PLANILLADO","FE_PLA_ANIOMES"]], use_container_width=True, hide_index=True)

    st.markdown("---"); st.markdown("### Seleccionar fila y corregir")
    opciones: dict[str, dict] = {}
    for f in filas:
        id_tramite = f.get("DIG_TRAMITE", "") or "(sin DIG_TRAMITE)"
        opciones[f"{id_tramite} | Cédula: {f.get('DIG_CEDULA','')} | Cobertura: {f.get('DIG_COBERTURA','')} | Mes: {f.get('FE_PLA_ANIOMES','')}"] = f

    seleccion = st.selectbox("Fila a corregir", options=list(opciones.keys()), key="select_fila_editar_cedulas")
    fila = opciones[seleccion]

    if not fila.get("DIG_TRAMITE"): st.error("Sin DIG_TRAMITE. No se permite modificar."); st.markdown("</div>", unsafe_allow_html=True); return
    if str(fila.get("DIG_PLANILLADO", "")).strip() != "S": st.error("Sin DIG_PLANILLADO='S'."); st.markdown("</div>", unsafe_allow_html=True); return

    fe_pla_fila = str(fila.get("FE_PLA_ANIOMES", "") or "").strip()
    if fe_pla_fila:
        conteo = _contar_por_fe_pla_aniomes(username, password, fe_pla_fila)
        if conteo.get("ok"):
            pct = conteo["pct"]
            color = "#4caf50" if pct >= 80 else ("#ff9800" if pct >= 50 else "#f44336")
            st.markdown(f"""<div style="background:#1e1e1e;border:1px solid #333;border-radius:8px;padding:10px 16px;margin-bottom:12px;">
            <span style="color:#aaa;font-size:0.85rem;">📊 Mes <b>{fe_pla_fila}</b></span><br>
            <span style="font-size:1.1rem;">✅ <b>{conteo['generados']}</b> / {conteo['total']} generados · ⏳ <b>{conteo['pendientes']}</b> pendientes</span>
            <div style="background:#333;border-radius:4px;height:8px;margin-top:6px;"><div style="background:{color};border-radius:4px;height:8px;width:{pct}%;"></div></div>
            </div>""", unsafe_allow_html=True)
        else:
            st.caption(f"⚠️ No se pudo consultar el mes {fe_pla_fila}")
        # Botón de descarga de errores
        with st.spinner("Preparando Excel de errores..."):
            excel_bytes = _generar_excel_errores_por_fe_pla(username, password, fe_pla_fila)
        if excel_bytes:
            st.download_button(
                label=f"📥 Descargar errores de {fe_pla_fila} (.xlsx)",
                data=excel_bytes,
                file_name=f"errores_{fe_pla_fila}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.caption(f"Sin errores registrados para {fe_pla_fila}")

    col1, col2 = st.columns(2)
    with col1:
        nueva_cedula = st.text_input("DIG_CEDULA titular", value=fila.get("DIG_CEDULA", ""), key="edit_cedula_titular")
        nueva_menor_edad = st.selectbox("DIG_MENOR_EDAD", options=["N", "S"], index=1 if str(fila.get("DIG_MENOR_EDAD","")).strip()=="S" else 0, key="edit_menor_edad")
    with col2:
        nuevo_dep1 = st.text_input("DIG_DEPENDIENTE_01", value=fila.get("DIG_DEPENDIENTE_01", ""), key="edit_dependiente_01")
        nuevo_dep2 = st.text_input("DIG_DEPENDIENTE_02", value=fila.get("DIG_DEPENDIENTE_02", ""), key="edit_dependiente_02")

    st.info("Resultado esperado: titular solo → CC.pdf; titular + dependientes → CC_01, CC_02, CC_03.")
    st.code(f"Trámite: {fila.get('DIG_TRAMITE')}\n1) Actualiza cédulas en Oracle\n2) DIG_COBERTURA='N'\n3) Regenera PDFs en {OUTPUT_ROOT}/{fila.get('DIG_TRAMITE')}\n4) Reemplaza SOLO CC*.pdf en repo\n5) No toca otros PDFs\n6) No crea carpetas", language="text")

    with st.form("form_guardar_regenerar_reemplazar_cc"):
        confirmar = st.checkbox("Confirmo corregir cédulas, regenerar cobertura y reemplazar solo CC*.pdf", key="confirmar_cedulas_regenerar_cc")
        frase = st.text_input("Escribir REGENERAR para confirmar", value="", key="frase_confirmacion_regenerar")
        submitted = st.form_submit_button("Guardar, regenerar y reemplazar SOLO CC*.pdf", use_container_width=True)

        if submitted:
            if not confirmar or frase.strip().upper() != "REGENERAR": st.warning("Falta confirmar y escribir REGENERAR."); st.markdown("</div>", unsafe_allow_html=True); return
            with st.spinner("Corrigiendo Oracle, regenerando PDFs y reemplazando solo CC*.pdf..."):
                try: resultado = _corregir_regenerar_y_reemplazar(username, password, fila, nueva_cedula, nueva_menor_edad, nuevo_dep1, nuevo_dep2)
                except Exception as exc: resultado = {"ok": False, "paso": "ERROR_NO_CONTROLADO", "error": str(exc)}

            if resultado.get("ok"):
                st.success("Corrección completada. CC*.pdf regenerados y reemplazados.")
                st.json({"destino": resultado.get("destino"), "pdfs_locales_generados": resultado.get("pdfs_locales_generados"),
                         "cc_eliminados_destino": resultado.get("reemplazo",{}).get("eliminados"),
                         "cc_copiados_destino": resultado.get("reemplazo",{}).get("copiados"),
                         "otros_no_tocados": resultado.get("reemplazo",{}).get("otros_no_tocados"),
                         "backup": resultado.get("reemplazo",{}).get("backup_dir"),
                         "manifest": resultado.get("reemplazo",{}).get("manifest_path")})
                st.session_state["cedulas_filas_encontradas"] = _buscar_por_tramite(username, password, str(fila.get("DIG_TRAMITE","")))
                st.rerun()
            else:
                st.error(f"No se completó. Paso: {resultado.get('paso')}")
                st.code(resultado.get("error", "Error desconocido"), language="text")
                st.json(resultado)
    st.markdown("</div>", unsafe_allow_html=True)
